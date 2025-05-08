from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from app.database import get_db
from app import crud, schemas
from typing import Optional
from app import models
import openpyxl
from app.models.equipment import Equipment
from app.models.category import Category
import io
from fastapi.responses import StreamingResponse
import pandas as pd
from app.database import get_db
from app.models.equipment import Equipment as EquipmentModel
from app.schemas.equipment import Equipment as EquipmentSchema
from math import exp

router = APIRouter(
    prefix="/equipments",
    tags=["Equipments"]
)

@router.get("/equipments/export", response_class=StreamingResponse)
def export_equipments_to_excel(db: Session = Depends(get_db)):
    equipments = db.query(Equipment).all()

    data = [{
        "name": eq.name,
        "warranty_years": eq.warranty_years,
        "min_temperature": eq.min_temperature,
        "max_temperature": eq.max_temperature,
        "link": eq.link,
        "category_id": eq.category_id,
        "mtbf_hours": eq.mtbf_hours,
    } for eq in equipments]

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=equipments_export.xlsx"}
    )

@router.post("/", response_model=schemas.Equipment)
def create_equipment(equipment: schemas.EquipmentCreate, db: Session = Depends(get_db)):
    # Проверка существования категории
    category = db.query(models.Category).filter(models.Category.id == equipment.category_id).first()
    if not category:
        raise HTTPException(status_code=400, detail=f"Категория с id {equipment.category_id} не найдена.")
    
    # Если всё ок, создаём оборудование
    db_equipment = crud.equipment.create_equipment(db, equipment)
    return schemas.Equipment.from_orm_with_calculations(db_equipment)

@router.get("/{equipment_id}", response_model=schemas.Equipment)
def read_equipment(equipment_id: int, db: Session = Depends(get_db)):
    db_equipment = crud.equipment.get_equipment(db, equipment_id=equipment_id)
    if db_equipment is None:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return schemas.Equipment.from_orm_with_calculations(db_equipment)


@router.get("/", response_model=list[schemas.Equipment])
def read_equipments(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    equipments = crud.equipment.get_equipments(db, skip=skip, limit=limit)
    return [schemas.Equipment.from_orm_with_calculations(eq) for eq in equipments]


@router.delete("/{equipment_id}", response_model=schemas.Equipment)
def delete_equipment(equipment_id: int, db: Session = Depends(get_db)):
    db_equipment = crud.equipment.delete_equipment(db, equipment_id)
    if db_equipment is None:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return db_equipment

@router.get("/{equipment_id}", response_model=schemas.Equipment)
def read_equipment(equipment_id: int, db: Session = Depends(get_db)):
    db_equipment = crud.equipment.get_equipment(db, equipment_id)
    if db_equipment is None:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return schemas.Equipment.from_orm_with_calculations(db_equipment)

@router.get("/", response_model=list[schemas.Equipment])
def read_equipments(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    db_equipments = crud.equipment.get_equipments(db, skip=skip, limit=limit)
    return [schemas.Equipment.from_orm_with_calculations(eq) for eq in db_equipments]

@router.put("/{equipment_id}", response_model=schemas.Equipment)
def update_equipment(equipment_id: int, equipment: schemas.EquipmentCreate, db: Session = Depends(get_db)):
    # Проверка существования категории
    category = db.query(models.Category).filter(models.Category.id == equipment.category_id).first()
    if not category:
        raise HTTPException(status_code=400, detail=f"Категория с id {equipment.category_id} не найдена.")

    db_equipment = crud.equipment.update_equipment(db, equipment_id, equipment)
    if db_equipment is None:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return schemas.Equipment.from_orm_with_calculations(db_equipment)


@router.get("/search/", response_model=list[schemas.Equipment])
def search_equipments(
    category_id: Optional[int] = None,
    min_temp: Optional[float] = None,
    max_temp: Optional[float] = None,
    min_mtbf: Optional[float] = None,
    name_contains: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    db_equipments = crud.equipment.search_equipments(
        db,
        category_id=category_id,
        min_temp=min_temp,
        max_temp=max_temp,
        min_mtbf=min_mtbf,
        name_contains=name_contains,
        skip=skip,
        limit=limit
    )
    return [schemas.Equipment.from_orm_with_calculations(eq) for eq in db_equipments]


@router.post("/import/")
def import_equipments_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате .xlsx")

    contents = file.file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    expected_headers = ['name', 'warranty_years', 'min_temperature', 'max_temperature', 'link', 'category_id', 'mtbf_hours']
    if headers != expected_headers:
        raise HTTPException(status_code=400, detail=f"Неверные заголовки. Ожидаются: {expected_headers}")

    success_count = 0
    errors = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        data = {header: cell.value for header, cell in zip(headers, row)}

        try:
            if not db.query(Category).filter_by(id=data["category_id"]).first():
                raise ValueError(f"Категория с id={data['category_id']} не найдена")

            equipment = Equipment(**data)
            db.add(equipment)
            success_count += 1
        except Exception as e:
            errors.append(f"Строка {idx}: {str(e)}")

    db.commit()

    return {
        "status": "Импорт завершён",
        "успешно добавлено": success_count,
        "ошибки": errors
    }


@router.get("/equipments/{equipment_id}/calculate_mtbf_exploitation/")
def calculate_mtbf_exploitation(
    equipment_id: int,
    temperature: float = Query(..., description="Температура эксплуатации в градусах Цельсия"),
    db: Session = Depends(get_db)
):
    db_equipment = db.query(EquipmentModel).filter(EquipmentModel.id == equipment_id).first()
    if not db_equipment:
        raise HTTPException(status_code=404, detail="Оборудование не найдено")

    # Получаем экземпляр схемы с расчётными полями
    equipment = EquipmentSchema.from_orm_with_calculations(db_equipment)

    # Используем уже имеющуюся функцию `calc_k`
    Ea = 0.28
    k_b = 8.617e-3
    try:
        T = temperature + 273
        term = (1 / T) - (1 / 298)
        exponent = (-Ea / k_b) * term
        Kt = round(exp(exponent), 5)
    except Exception:
        raise HTTPException(status_code=400, detail="Невозможно вычислить коэффициент K")

    # Вычисляем интенсивность и MTBF
    if not equipment.mtbf_hours:
        raise HTTPException(status_code=400, detail="Отсутствует базовый MTBF")

    lambda_e = 1 / equipment.mtbf_hours
    lambda_ex = lambda_e * Kt
    mtbf_ex = 1 / lambda_ex

    return {
        'id': equipment.id,
        "lambda_exploitation": round(lambda_ex, 6),
        "mtbf_exploitation": round(mtbf_ex, 3)
    }